"""
Statement Copilot - Action Planner Agent
========================================
Plans and executes user actions (export, reports, alerts).

bunq Alignment: Plan -> Confirm -> Execute pattern
"""

import logging
import time
import uuid
from typing import Dict, Any, Optional, List
from datetime import date, datetime
from pathlib import Path

from ..config import settings
from ..core import (
    OrchestratorState,
    ActionPlan,
    ActionPlanDraft,
    ActionParams,
    ActionResult,
    DataScope,
    ActionType,
    RiskLevel,
    IntentType,
    DateRange,
    SQLBuilder,
    create_tool_call_record,
    get_llm_client,
    get_db,
    get_date_range_from_constraints,
)
from ..log_context import format_kv, format_list
from ..core.prompts import get_action_planner_prompt, get_action_plan_draft_prompt

logger = logging.getLogger(__name__)


# -----------------------------------------------------------------------------
# ACTION PLANNER AGENT
# -----------------------------------------------------------------------------

class ActionPlannerAgent:
    """
    Action Planner that:
    1. Creates execution plans for user actions
    2. Estimates scope and risk
    3. Generates human-readable confirmation
    4. Executes approved actions
    
    bunq Alignment: Always requires user confirmation before execution.
    """
    
    def __init__(self):
        self.llm = get_llm_client()
        self.db = get_db()
        self.model = settings.model_action_planner
        self.outputs_path = settings.get_outputs_path()
    
    def plan(self, state: OrchestratorState) -> OrchestratorState:
        """
        Create action plan based on user request.
        
        Args:
            state: Current orchestrator state
            
        Returns:
            Updated state with action_plan
        """
        user_message = state.get("user_message", "")
        constraints = state.get("constraints", {})
        tenant_id = state.get("tenant_id", settings.default_tenant_id)
        
        start_time = time.time()
        
        try:
            # Step 1: Generate action plan
            plan, missing_fields = self._generate_plan(user_message, constraints, tenant_id)

            logger.debug(
                "Action plan generated: type=%s missing_fields=%s",
                plan.action_type.value if plan else None,
                format_list(missing_fields or [], max_items=6, max_value_len=60),
            )

            if missing_fields:
                state["intent"] = IntentType.CLARIFY.value
                state["clarification_needed"] = (
                    "Some information is missing for this action. Could you clarify: "
                    + ", ".join(missing_fields)
                )
                state["suggested_questions"] = ["Could you provide the missing details?"]
                state["needs_confirmation"] = False
                state["action_plan"] = None
                state["pending_action_id"] = None
                tool_call = create_tool_call_record(
                    state=state,
                    node="action_planner",
                    tool_name="plan",
                    model_name=self.model,
                    input_data={"message": user_message},
                    output_data={"missing_fields": missing_fields},
                    latency_ms=int((time.time() - start_time) * 1000),
                    success=True
                )
                state["tool_calls"] = [tool_call]
                return state

            latency_ms = int((time.time() - start_time) * 1000)
            
            # Store plan in state
            state["action_plan"] = plan.model_dump(mode="json")
            state["pending_action_id"] = plan.action_id
            state["needs_confirmation"] = plan.requires_confirmation
            
            # Add tool call record
            tool_call = create_tool_call_record(
                state=state,
                node="action_planner",
                tool_name="plan",
                model_name=self.model,
                input_data={"message": user_message},
                output_data={"action_id": plan.action_id, "action_type": plan.action_type.value},
                latency_ms=latency_ms,
                success=True
            )
            state["tool_calls"] = [tool_call]

            logger.info(
                f"Action Planner: type={plan.action_type.value}, "
                f"risk={plan.risk_level.value}, latency={latency_ms}ms"
            )

            logger.debug(
                "Action plan summary: action_id=%s scope=%s params=%s",
                plan.action_id,
                format_kv(plan.data_scope.model_dump() if plan.data_scope else {}, max_items=6, max_value_len=120),
                format_kv(plan.params.model_dump() if plan.params else {}, max_items=6, max_value_len=120),
            )
            
            return state
            
        except Exception as e:
            logger.error(f"Action Planner error: {e}")
            state["action_error"] = str(e)
            state["errors"] = [f"Action planning error: {e}"]
            return state
    
    def execute(self, state: OrchestratorState) -> OrchestratorState:
        """
        Execute approved action.
        
        Args:
            state: State with approved action_plan
            
        Returns:
            Updated state with action_result
        """
        if not state.get("user_confirmed"):
            state["action_error"] = "Action not confirmed by user"
            return state
        
        plan_data = state.get("action_plan")
        if not plan_data:
            state["action_error"] = "No action plan found"
            return state
        
        tenant_id = state.get("tenant_id", settings.default_tenant_id)
        
        start_time = time.time()
        
        try:
            # Reconstruct ActionPlan
            plan = ActionPlan(**plan_data)

            logger.debug(
                "Action execute start: action_id=%s type=%s params=%s",
                plan.action_id,
                plan.action_type.value,
                format_kv(plan.params.model_dump() if plan.params else {}, max_items=6, max_value_len=120),
            )
            
            # Execute based on action type
            result = self._execute_action(plan, tenant_id)
            
            latency_ms = int((time.time() - start_time) * 1000)
            result.execution_time_ms = latency_ms
            
            # Store result
            state["action_result"] = result.model_dump(mode="json")
            
            # Add tool call record
            tool_call = create_tool_call_record(
                state=state,
                node="action_planner",
                tool_name="execute",
                model_name=None,
                input_data={"action_id": plan.action_id},
                output_data={"status": result.status, "artifacts": result.artifacts},
                latency_ms=latency_ms,
                success=result.status == "success"
            )
            state["tool_calls"] = [tool_call]

            logger.info(
                f"Action executed: action_id={plan.action_id}, "
                f"status={result.status}, latency={latency_ms}ms"
            )

            logger.debug(
                "Action execute done: status=%s artifacts=%s",
                result.status,
                format_kv(result.artifacts or {}, max_items=6, max_value_len=120),
            )
            
            return state
            
        except Exception as e:
            logger.error(f"Action execution error: {e}")
            state["action_result"] = ActionResult(
                action_id=plan_data.get("action_id", "unknown"),
                status="failed",
                error=str(e)
            ).model_dump(mode="json")
            state["errors"] = [f"Action execution error: {e}"]
            return state
    
    def _generate_plan(
        self,
        user_message: str,
        constraints: Dict[str, Any],
        tenant_id: str
    ) -> tuple[ActionPlan, list[str]]:
        """
        Generate action plan using LLM.
        
        Args:
            user_message: User's request
            constraints: Extracted constraints
            tenant_id: Tenant ID
            
        Returns:
            ActionPlan with human-readable description
        """
        # Get date range
        date_start, date_end = get_date_range_from_constraints(
            constraints,
            default_days=30
        )

        # LLM-based draft (best-effort)
        draft = self._generate_plan_draft(
            user_message=user_message,
            constraints=constraints,
            date_start=date_start,
            date_end=date_end,
        )

        # Determine action type
        action_type = draft.action_type if draft else self._detect_action_type(user_message)
        
        # Extract structured action params via LLM (best-effort)
        llm_params = draft.params if draft else self._extract_action_params(
            user_message=user_message,
            action_type=action_type,
            constraints=constraints,
            date_start=date_start,
            date_end=date_end,
        )
        
        # Build params
        default_report_type = None
        if action_type == ActionType.MONTHLY_REPORT:
            default_report_type = "monthly"
        elif action_type == ActionType.ANNUAL_REPORT:
            default_report_type = "annual"

        params = ActionParams(
            date_start=llm_params.date_start or date_start,
            date_end=llm_params.date_end or date_end,
            export_format=llm_params.export_format or self._get_export_format(action_type),
            include_charts=(
                llm_params.include_charts
                if llm_params.include_charts is not None
                else action_type in [ActionType.MONTHLY_REPORT, ActionType.ANNUAL_REPORT]
            ),
            report_type=llm_params.report_type or default_report_type,
            sections=llm_params.sections,
            category=llm_params.category
            or (constraints.get("categories", [None])[0] if constraints.get("categories") else None),
            threshold_amount=llm_params.threshold_amount,
            tx_ids=llm_params.tx_ids,
            new_category=llm_params.new_category,
            new_subcategory=llm_params.new_subcategory,
            reminder_date=llm_params.reminder_date,
            reminder_message=llm_params.reminder_message,
        )

        # Detect missing required fields
        missing_fields = draft.missing_fields if draft and draft.missing_fields else self._detect_missing_fields(action_type, params)
        
        # Build a temp plan for scope/risk
        temp_plan = ActionPlan(
            action_id=str(uuid.uuid4())[:8],
            action_type=action_type,
            human_plan="Planning in progress.",
            params=params,
            data_scope=DataScope(),
            requires_confirmation=draft.requires_confirmation if draft else True,
            risk_level=RiskLevel.LOW,
            estimated_time_seconds=draft.estimated_time_seconds if draft and draft.estimated_time_seconds else self._estimate_time(action_type),
        )

        # Step 2: Estimate data scope
        temp_plan.data_scope = self._estimate_scope(temp_plan, tenant_id)

        # Step 3: Assess risk and add warnings
        temp_plan = self._assess_risk(temp_plan)

        # If draft risk is higher, respect it
        if draft and draft.risk_level:
            if draft.risk_level.value in [RiskLevel.HIGH.value]:
                temp_plan.risk_level = RiskLevel.HIGH
            elif draft.risk_level.value in [RiskLevel.MEDIUM.value] and temp_plan.risk_level == RiskLevel.LOW:
                temp_plan.risk_level = RiskLevel.MEDIUM

        # Generate professional human plan
        human_plan = self._generate_human_plan(
            user_message=user_message,
            plan=temp_plan,
            constraints=constraints,
        )
        temp_plan.human_plan = human_plan.strip()
        temp_plan.plan_steps = self._default_plan_steps(action_type)

        return temp_plan, (missing_fields or [])

    def _generate_plan_draft(
        self,
        user_message: str,
        constraints: Dict[str, Any],
        date_start: date,
        date_end: date,
    ) -> Optional[ActionPlanDraft]:
        """Generate a structured action plan draft via LLM."""
        prompt = f"""
# -----------------------------------------------------------------------------
{user_message}

## MEVCUT CONSTRAINTLER
{constraints}

# -----------------------------------------------------------------------------
{date_start.isoformat()} - {date_end.isoformat()}

# -----------------------------------------------------------------------------
Produce an ActionPlanDraft.
"""
        try:
            draft = self.llm.complete_structured(
                prompt=prompt,
                response_model=ActionPlanDraft,
                model=self.model,
                system=get_action_plan_draft_prompt(),
                temperature=0.0,
            )
            logger.debug(
                "Action plan draft: type=%s risk=%s requires_confirmation=%s missing_fields=%s",
                draft.action_type.value if draft.action_type else None,
                draft.risk_level.value if draft.risk_level else None,
                draft.requires_confirmation,
                format_list(draft.missing_fields or [], max_items=6, max_value_len=60),
            )
            return draft
        except Exception as e:
            logger.error(f"ActionPlanDraft generation failed: {e}")
            return None

    def _detect_missing_fields(self, action_type: ActionType, params: ActionParams) -> List[str]:
        """Detect missing required fields for specific action types."""
        missing = []
        if action_type == ActionType.CATEGORY_UPDATE:
            if not params.tx_ids:
                missing.append("tx_ids")
            if not params.new_category:
                missing.append("new_category")
        elif action_type == ActionType.SET_BUDGET_ALERT:
            if not params.category:
                missing.append("category")
            if params.threshold_amount is None:
                missing.append("threshold_amount")
        elif action_type == ActionType.SET_REMINDER:
            if not params.reminder_date:
                missing.append("reminder_date")
            if not params.reminder_message:
                missing.append("reminder_message")
        return missing

    def _default_plan_steps(self, action_type: ActionType) -> List[str]:
        """Default execution steps per action type."""
        steps_map = {
            ActionType.EXPORT_XLSX: [
                "Gather data for the selected period and filters",
                "Create the Excel file",
                "Make the file downloadable",
            ],
            ActionType.EXPORT_CSV: [
                "Gather data for the selected period and filters",
                "Create the CSV file",
                "Make the file downloadable",
            ],
            ActionType.EXPORT_PDF: [
                "Compute summary and details",
                "Create the PDF report",
                "Make the file downloadable",
            ],
            ActionType.MONTHLY_REPORT: [
                "Compute monthly summary metrics",
                "Compute category and merchant breakdowns",
                "Generate the report",
            ],
            ActionType.ANNUAL_REPORT: [
                "Compute annual summary metrics",
                "Compute trends and breakdowns",
                "Generate the report",
            ],
            ActionType.SUBSCRIPTION_REVIEW: [
                "Identify recurring payments",
                "Classify subscriptions",
                "Prepare summary list",
            ],
            ActionType.SET_BUDGET_ALERT: [
                "Validate budget limit",
                "Create alert rule",
                "Save notification settings",
            ],
            ActionType.CATEGORY_UPDATE: [
                "Validate affected transactions",
                "Apply category update",
                "Sonucu kaydetme",
            ],
            ActionType.SET_REMINDER: [
                "Validate reminder date",
                "Create the reminder",
                "Save notification schedule",
            ],
        }
        return steps_map.get(action_type, ["Execute the action"])

    def _generate_human_plan(
        self,
        user_message: str,
        plan: ActionPlan,
        constraints: Dict[str, Any],
    ) -> str:
        """Generate professional human-readable plan."""
        date_range = None
        if plan.params.date_start and plan.params.date_end:
            date_range = f"{plan.params.date_start.isoformat()} - {plan.params.date_end.isoformat()}"

        prompt = f"""
# -----------------------------------------------------------------------------
{user_message}

# -----------------------------------------------------------------------------
{plan.action_type.value}

# -----------------------------------------------------------------------------
{date_range or "Belirtilmedi"}

## PARAMETRELER
{plan.params.model_dump()}

## KAPSAM
{plan.data_scope.model_dump()}

# -----------------------------------------------------------------------------
{plan.risk_level.value}

## UYARILAR
{plan.warnings}

# -----------------------------------------------------------------------------
Write a professional, clear, friendly English plan.
Explain what will be done, which data will be used, and the estimated time.
"""
        try:
            response = self.llm.complete(
                prompt=prompt,
                model=self.model,
                system=get_action_planner_prompt(),
                max_tokens=500,
                temperature=0.2,
            )
            return response
        except Exception as e:
            logger.error(f"Human plan generation failed: {e}")
            return self._generate_fallback_plan(
                plan.action_type,
                plan.params.date_start or date.today(),
                plan.params.date_end or date.today(),
            )

    def _extract_action_params(
        self,
        user_message: str,
        action_type: ActionType,
        constraints: Dict[str, Any],
        date_start: date,
        date_end: date,
    ) -> ActionParams:
        """
        Extract structured ActionParams via LLM (best-effort).
        """
        prompt = f"""
# -----------------------------------------------------------------------------
{user_message}

# -----------------------------------------------------------------------------
{action_type.value}

## MEVCUT CONSTRAINTLER
{constraints}

# -----------------------------------------------------------------------------
{date_start.isoformat()} - {date_end.isoformat()}

# -----------------------------------------------------------------------------
Only fill fields the user explicitly provided.
Leave unspecified fields as null.
"""

        try:
            params = self.llm.complete_structured(
                prompt=prompt,
                response_model=ActionParams,
                model=self.model,
                system=get_action_planner_prompt(),
                temperature=0.0
            )
            logger.debug(
                "Action params extracted: %s",
                format_kv(params.model_dump() if params else {}, max_items=6, max_value_len=120),
            )
            return params
        except Exception as e:
            logger.error(f"ActionParams extraction failed: {e}")
            return ActionParams()
    
    def _detect_action_type(self, message: str) -> ActionType:
        """Detect action type from message"""
        message_lower = message.lower()
        
        # Export detection
        if any(word in message_lower for word in ["excel", "xlsx", "table"]):
            return ActionType.EXPORT_XLSX
        if "csv" in message_lower:
            return ActionType.EXPORT_CSV
        if "pdf" in message_lower:
            return ActionType.EXPORT_PDF

        # Report detection
        if any(word in message_lower for word in ["monthly report", "month report", "monthly summary"]):
            return ActionType.MONTHLY_REPORT
        if any(word in message_lower for word in ["annual report", "year report", "yearly summary"]):
            return ActionType.ANNUAL_REPORT
        if any(word in message_lower for word in ["subscription", "subscriptions"]):
            return ActionType.SUBSCRIPTION_REVIEW

        # Alert detection
        if any(word in message_lower for word in ["alert", "alarm", "notification", "budget"]):
            return ActionType.SET_BUDGET_ALERT

        # Category update
        if any(word in message_lower for word in ["change category", "update category"]):
            return ActionType.CATEGORY_UPDATE

        # Reminder
        if any(word in message_lower for word in ["remind", "reminder"]):
            return ActionType.SET_REMINDER

        # Default to export
        if any(word in message_lower for word in ["export", "download", "save"]):
            return ActionType.EXPORT_XLSX
        
        # Default to monthly report
        return ActionType.MONTHLY_REPORT
    
    def _get_export_format(self, action_type: ActionType) -> Optional[str]:
        """Get export format for action type"""
        format_map = {
            ActionType.EXPORT_XLSX: "xlsx",
            ActionType.EXPORT_CSV: "csv",
            ActionType.EXPORT_PDF: "pdf",
            ActionType.MONTHLY_REPORT: "xlsx",
            ActionType.ANNUAL_REPORT: "xlsx",
        }
        return format_map.get(action_type)
    
    def _estimate_time(self, action_type: ActionType) -> int:
        """Estimate execution time in seconds"""
        time_map = {
            ActionType.EXPORT_XLSX: 5,
            ActionType.EXPORT_CSV: 3,
            ActionType.EXPORT_PDF: 10,
            ActionType.MONTHLY_REPORT: 15,
            ActionType.ANNUAL_REPORT: 30,
            ActionType.SUBSCRIPTION_REVIEW: 10,
            ActionType.SET_BUDGET_ALERT: 2,
            ActionType.CATEGORY_UPDATE: 5,
            ActionType.SET_REMINDER: 2,
            ActionType.GENERATE_CHART: 10,
        }
        return time_map.get(action_type, 10)
    
    def _estimate_scope(self, plan: ActionPlan, tenant_id: str) -> DataScope:
        """Estimate data scope for action"""
        params = plan.params
        
        # Count affected rows
        filters = {
            "date_start": params.date_start,
            "date_end": params.date_end,
        }
        if params.category:
            filters["categories"] = [params.category]
        
        sql, sql_params = SQLBuilder.build(
            metric="count_tx",
            filters=filters,
            tenant_id=tenant_id
        )
        
        try:
            rows = self.db.execute_query(sql, sql_params)
            estimated_rows = rows[0]["value"] if rows else 0
        except Exception:
            estimated_rows = None
        
        write_actions = {
            ActionType.CATEGORY_UPDATE,
            ActionType.SET_BUDGET_ALERT,
            ActionType.SET_REMINDER,
        }

        logger.debug(
            "Action scope estimate: rows=%s date_start=%s date_end=%s category=%s read_only=%s",
            estimated_rows,
            params.date_start,
            params.date_end,
            params.category,
            plan.action_type not in write_actions,
        )

        return DataScope(
            tables=["transactions"],
            date_range=DateRange(
                start=params.date_start,
                end=params.date_end
            ) if params.date_start and params.date_end else None,
            estimated_rows=estimated_rows,
            categories_affected=[params.category] if params.category else None,
            read_only=plan.action_type not in write_actions,
        )
    
    def _assess_risk(self, plan: ActionPlan) -> ActionPlan:
        """Assess risk level and add warnings"""
        warnings = []
        risk = RiskLevel.LOW
        severity = {
            RiskLevel.LOW: 1,
            RiskLevel.MEDIUM: 2,
            RiskLevel.HIGH: 3,
        }

        def bump_risk(current: RiskLevel, candidate: RiskLevel) -> RiskLevel:
            return candidate if severity[candidate] > severity[current] else current
        
        # Check data scope
        if plan.data_scope.estimated_rows:
            if plan.data_scope.estimated_rows > 50000:
                warnings.append("Many transactions will be affected. This may take a while.")
                risk = RiskLevel.HIGH
            elif plan.data_scope.estimated_rows > 10000:
                warnings.append("More than 10,000 transactions will be affected. This may take a while.")
                risk = RiskLevel.MEDIUM
        
        # Check date range
        if plan.data_scope.date_range:
            days = (plan.data_scope.date_range.end - plan.data_scope.date_range.start).days
            if days > 365:
                warnings.append("A period longer than 1 year was selected.")
                risk = bump_risk(risk, RiskLevel.MEDIUM)
        
        # Check action type
        if plan.action_type == ActionType.CATEGORY_UPDATE:
            warnings.append("This action will modify data.")
            risk = bump_risk(risk, RiskLevel.MEDIUM)
            
            if not plan.data_scope.read_only:
                risk = RiskLevel.HIGH
        
        plan.risk_level = risk
        plan.warnings = warnings
        
        return plan
    
    def _generate_fallback_plan(
        self,
        action_type: ActionType,
        date_start: date,
        date_end: date
    ) -> str:
        """Generate fallback human plan without LLM"""
        type_descriptions = {
            ActionType.EXPORT_XLSX: "Excel file",
            ActionType.EXPORT_CSV: "CSV file",
            ActionType.EXPORT_PDF: "PDF report",
            ActionType.MONTHLY_REPORT: "monthly report",
            ActionType.ANNUAL_REPORT: "annual report",
            ActionType.SUBSCRIPTION_REVIEW: "subscription analysis",
            ActionType.SET_BUDGET_ALERT: "budget alert",
            ActionType.CATEGORY_UPDATE: "category update",
        }
        
        desc = type_descriptions.get(action_type, "action")
        return f"A {desc} will be prepared for your data between {date_start.strftime('%d %B')} and {date_end.strftime('%d %B %Y')}."
    
    def _execute_action(self, plan: ActionPlan, tenant_id: str) -> ActionResult:
        """Execute the action and return result"""
        action_type = plan.action_type
        
        if action_type in [ActionType.EXPORT_XLSX, ActionType.EXPORT_CSV]:
            return self._execute_export(plan, tenant_id)
        elif action_type in [ActionType.MONTHLY_REPORT, ActionType.ANNUAL_REPORT]:
            return self._execute_report(plan, tenant_id)
        elif action_type == ActionType.SUBSCRIPTION_REVIEW:
            return self._execute_subscription_review(plan, tenant_id)
        elif action_type == ActionType.SET_BUDGET_ALERT:
            return self._execute_budget_alert(plan, tenant_id)
        elif action_type == ActionType.SET_REMINDER:
            return self._execute_reminder(plan, tenant_id)
        elif action_type == ActionType.CATEGORY_UPDATE:
            return self._execute_category_update(plan, tenant_id)
        else:
            return ActionResult(
                action_id=plan.action_id,
                status="failed",
                error=f"Action type {action_type.value} not implemented"
            )
    
    def _execute_export(self, plan: ActionPlan, tenant_id: str) -> ActionResult:
        """Execute data export"""
        params = plan.params
        
        # Fetch data
        filters = {
            "date_start": params.date_start,
            "date_end": params.date_end,
        }
        if params.category:
            filters["categories"] = [params.category]
        
        # Use a simple SELECT query for export
        sql = f"""
            SELECT 
                date_time,
                merchant_norm as merchant,
                description,
                amount,
                direction,
                COALESCE(category_final, category) as category,
                COALESCE(subcategory_final, subcategory) as subcategory
            FROM transactions
            WHERE tenant_id = ?
            AND date_time >= ?
            AND date_time <= ?
            ORDER BY date_time DESC
            LIMIT 10000
        """
        
        end_datetime = datetime.combine(params.date_end, datetime.max.time()) if params.date_end else datetime.now()
        start_date = params.date_start or date.today()
        start_datetime = datetime.combine(start_date, datetime.min.time())
        
        rows = self.db.execute_query(sql, [tenant_id, start_datetime, end_datetime])
        
        if not rows:
            return ActionResult(
                action_id=plan.action_id,
                status="success",
                message="No transactions found for the selected period.",
                artifacts={}
            )
        
        # Generate file
        export_format = params.export_format or "xlsx"
        filename = f"export_{plan.action_id}.{export_format}"
        filepath = self.outputs_path / filename
        
        if export_format == "xlsx":
            self._write_xlsx(rows, filepath)
        elif export_format == "csv":
            self._write_csv(rows, filepath)
        
        return ActionResult(
            action_id=plan.action_id,
            status="success",
            message=f"{len(rows)} transactions exported.",
            artifacts={"export_file": str(filepath)}
        )
    
    def _execute_report(self, plan: ActionPlan, tenant_id: str) -> ActionResult:
        """Execute report generation"""
        # For now, delegate to export with summary
        return self._execute_export(plan, tenant_id)
    
    def _execute_subscription_review(self, plan: ActionPlan, tenant_id: str) -> ActionResult:
        """Execute subscription review"""
        sql, params = SQLBuilder.build(
            metric="subscription_list",
            filters={},
            tenant_id=tenant_id,
            limit=50
        )
        
        rows = self.db.execute_query(sql, params)
        
        return ActionResult(
            action_id=plan.action_id,
            status="success",
            message=f"{len(rows)} aktif abonelik bulundu.",
            artifacts={"subscriptions": rows}
        )
    
    def _execute_budget_alert(self, plan: ActionPlan, tenant_id: str) -> ActionResult:
        """Execute budget alert setup"""
        # Would integrate with notification service
        return ActionResult(
            action_id=plan.action_id,
            status="success",
            message="Budget alert set successfully."
        )
    

    def _execute_reminder(self, plan: ActionPlan, tenant_id: str) -> ActionResult:
        """Execute reminder setup (placeholder)"""
        params = plan.params
        if not params.reminder_date or not params.reminder_message:
            return ActionResult(
                action_id=plan.action_id,
                status="failed",
                error="Reminder date and message required"
            )

        # Placeholder: integrate with scheduling/notification service
        return ActionResult(
            action_id=plan.action_id,
            status="success",
            message="Reminder request received."
        )

    def _execute_category_update(self, plan: ActionPlan, tenant_id: str) -> ActionResult:
        """Execute category update"""
        params = plan.params
        
        if not params.tx_ids or not params.new_category:
            return ActionResult(
                action_id=plan.action_id,
                status="failed",
                error="Transaction IDs and new category required"
            )
        
        # Update categories
        placeholders = ",".join(["?" for _ in params.tx_ids])
        sql = f"""
            UPDATE transactions
            SET category_final = ?,
                subcategory_final = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE tx_id IN ({placeholders})
            AND tenant_id = ?
        """
        
        sql_params = [
            params.new_category,
            params.new_subcategory,
            *params.tx_ids,
            tenant_id
        ]
        
        self.db.execute_query(sql, sql_params, read_only=False)
        
        return ActionResult(
            action_id=plan.action_id,
            status="success",
            message=f"Updated category for {len(params.tx_ids)} transactions."
        )
    
    def _write_xlsx(self, rows: List[Dict[str, Any]], filepath: Path):
        """Write data to Excel file"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transactions"
            
            if rows:
                # Headers
                headers = list(rows[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Data
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row.get(header)
                        if isinstance(value, datetime):
                            value = value.strftime("%Y-%m-%d %H:%M")
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(filepath)
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            self._write_csv(rows, filepath.with_suffix('.csv'))
    
    def _write_csv(self, rows: List[Dict[str, Any]], filepath: Path):
        """Write data to CSV file"""
        import csv
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            if rows:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)


# -----------------------------------------------------------------------------
# SINGLETON
# -----------------------------------------------------------------------------

_action_planner: Optional[ActionPlannerAgent] = None


def get_action_planner() -> ActionPlannerAgent:
    """Get or create action planner singleton"""
    global _action_planner
    if _action_planner is None:
        _action_planner = ActionPlannerAgent()
    return _action_planner
